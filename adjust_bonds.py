import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from common import *
import excel2img

request_headers_file = 'request_headers.txt'
# 读取文件内容
request_headers = read_jisilu_request_headers_file(request_headers_file)

# 输出文件内容
print(request_headers)


def fetch_other_bonds():
    #下修相关内容
    url = "https://www.jisilu.cn/webapi/cb/list/"

    # 发起 HTTP GET 请求
    #my_header = request_headers
    my_header = {
    'cookie': request_headers["cookie"],
    'Referer': 'https://www.jisilu.cn/web/data/cb/list', 
    'Accept': 'application/json, text/plain, */*',
    #':authority':'www.jisilu.cn',
    #':method':'GET',
    #':path':'/webapi/cb/list/',
    #':scheme':'https',
    'Accept-Encoding':'gzip, deflate, br',
    'Accept-Language':'en,zh-CN;q=0.9,zh;q=0.8,zh-HK;q=0.7,zh-TW;q=0.6,ko;q=0.5',
    'Columns':'1,70,2,3,5,6,7,8,9,10,11,12,14,15,77,78,79,80,81,82,83,84,16,19,22,23,24,72,25,26,27,28,29,30,31,32,33,34,35,75,44,46,47,50,52,53,74,73,54,55,56,57,58,59,60,61,62,76,63,67,71',
    'Dnt':'1',
    'If-Modified-Since':'Mon, 04 Sep 2023 11:10:38 GMT',
    'Init':'1',
    'Referer':'https://www.jisilu.cn/web/data/cb/list',
    'Sec-Ch-Ua':'"Chromium";v="116", "Not)A;Brand";v="24", "Google Chrome";v="116"',
    'Sec-Ch-Ua-Mobile':'?1',
    'Sec-Ch-Ua-Platform':"Android",
    'Sec-Fetch-Dest':'empty',
    'Sec-Fetch-Mode':'cors',
    'Sec-Fetch-Site':'same-origin',
    'User-Agent':'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36'
    }
    #my_header['Referer'] = 'https://www.jisilu.cn/web/data/cb/list'
    response = requests.get(url,headers=my_header)

    if response.status_code == 200:
        data = response.json()
        return data.get("data", [])
    else:
        print(f"请求失败，状态码: {response.status_code}")
        return []

def fetch_all_convertible_bonds():
    #下修相关内容
    url = "https://www.jisilu.cn/webapi/cb/adjust/"

    # 发起 HTTP GET 请求
    response = requests.get(url,headers=request_headers)

    if response.status_code == 200:
        data = response.json()
        return data.get("data", [])
    else:
        print(f"请求失败，状态码: {response.status_code}")
        return []

# 抓取所有转债数据
all_bonds_data = fetch_all_convertible_bonds()

other_bonds = fetch_other_bonds()


excel_header =   ['转债名称','转债代码','收盘价','溢价率','下修日计数','到期税前收益率','剩余年限','备注']
# 创建字体对象
title_font = Font(name='楷体')
header_font = Font(name='微软雅黑',bold=False)
backup_font = Font(name='微软雅黑',size=9)
title_height = 30
header_fill = PatternFill(fill_type='solid', fgColor='f2f2f2')


# 创建一个新的工作簿
workbook = Workbook()
sheet = workbook.active
sheet.merge_cells('A1:F1')
sheet['A1'] = '即将下修转债列表'
sheet['A1'].font = title_font
sheet.row_dimensions[1].height = 22
sheet.column_dimensions['A'].width = 12
sheet.column_dimensions['E'].width = 14
sheet.column_dimensions['F'].width = 8
sheet.column_dimensions['H'].width = 24

sheet.append(excel_header)
for cell in sheet[2]:
        #if(cell.internal_value != "代码"):
        cell.fill = header_fill
        cell.font = header_font
index = 2
select_stocks=[]
tip_str = "[即将满足下修条件]:"
for bond_data in all_bonds_data:
    bond_code = bond_data["bond_id"]
    adjust_count = find_property_value(all_bonds_data,bond_code, 'adjust_count')
    print(bond_code)
    print(bond_data)
    #print(adjust_count)
    adjust = parse_adjust_string(adjust_count)
    if adjust == None:
        continue
    if adjust[0] == 0:
        continue
    if adjust[0] > 15:
        continue

    #再判断收盘价，如果收盘价高于122，就不统计了
    price = find_property_value(all_bonds_data,bond_code, 'price')
    if price>122:
        continue

    index = index + 1
    bond_nm = find_property_value(all_bonds_data,bond_code, 'bond_nm')
    bond_id = find_property_value(all_bonds_data,bond_code, 'bond_id')
    #bond_id = int(bond_id)
    #price = find_property_value(all_bonds_data,bond_code, 'price')
    premium_rt = find_property_value(all_bonds_data,bond_code, 'premium_rt')
    premium_rt = premium_rt/100.00

    #税前收益率,需要从other bonds里面获取
    ytm_rt = find_property_value(other_bonds,bond_code, 'ytm_rt')

    #剩余到期年限
    year_left = find_property_value(other_bonds,bond_code, 'year_left')

    #premium_rt_str =  str(premium_rt) + '%'
    readjust_dt = find_property_value(all_bonds_data,bond_code, 'readjust_dt')
    bond_backup = find_backup_content_by_id(bond_id)
    item = [bond_nm,int(bond_id),price,premium_rt,adjust_count,f"{ytm_rt}%",f"{year_left}",bond_backup]
    sheet.append(item)
    stock_id = find_property_value(all_bonds_data,bond_code, 'stock_id')
    select_stocks.append(stock_id)
    #如果只剩下3天了，那么标红，做醒目处理
    red_font  = Font(color='FF4500')  
    if adjust[1]-adjust[0]<4:
        pos = 'E'+str(index)
        sheet[pos].font = red_font
        item_str = bond_nm + "(最少"+str(adjust[1]-adjust[0])+"天),"
        tip_str = tip_str + item_str
    
    #如果到期收益率是正的，那么标红，做醒目处理
    if ytm_rt>0:
         pos = 'F'+str(index)
         sheet[pos].font = red_font


    

#设置百分比
for cell in sheet['D']:
    pos = 'D'+str(cell.row)
    print(sheet[pos].value)
    if(is_number(sheet[pos].value)):
        cell.number_format = '0.00%'

blue_font  = Font(color='0000FF')  # 蓝色字体
for cell in sheet['B']:
    pos = 'B'+str(cell.row)
    print(sheet[pos].value)
    if(is_integer(sheet[pos].value)):
        cell.font = blue_font

yahei_font  = Font(name='微软雅黑',bold=False)
for cell in sheet['A']:
    pos = 'A'+str(cell.row)
    #print(sheet[pos].value)
    #if(is_integer(sheet[pos].value)):
    cell.font = yahei_font

#设置备注文字样式
for cell in sheet['H']:
    pos = 'H'+str(cell.row)
    sheet[pos].font = backup_font
    #sheet[pos].alignment = Alignment(wrap_text=True)
             
#设置居中
# TOdo自动调整宽度
for row in sheet.iter_rows(min_row=1, max_row=index, min_col=1, max_col=8):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

# 保存stock_id
stock_id_list_file = get_file_path("即将下修转债列表-stock_id.txt")
write_array_to_file(select_stocks,stock_id_list_file)

# 保存工作簿
adjust_list_file = get_file_path("即将下修转债列表.xlsx")
workbook.save(adjust_list_file)

adjust_list_image_file = os.path.join(get_image_path(),'即将下修转债列表.png')
excel2img.export_img(adjust_list_file, adjust_list_image_file, "Sheet", None)

append_reminder(tip_str)

