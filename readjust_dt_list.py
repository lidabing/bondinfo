#下修重算起始日即将到期的转载
#价格在110以内
import requests
import re
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from common import *
from common_net import *
import excel2img

    
# 抓取所有转债数据
adjust_bonds = fetch_adjust_bonds_info()
base_bonds = fetch_base_bonds_info()

excel_header =   ['转债名称','转债代码','收盘价','溢价率','下修重算日','到期税前收益率','剩余年限','备注']
# 创建字体对象
title_font = Font(name='微软雅黑')
header_font = Font(name='微软雅黑',bold=False)
backup_font = Font(name='微软雅黑',size=9)
title_height = 30
header_fill = PatternFill(fill_type='solid', fgColor='f2f2f2')

# 创建一个新的工作簿
workbook = Workbook()
sheet = workbook.active
sheet.merge_cells('A1:F1')
sheet['A1'] = '下修重算日即将到期转债'
sheet['A1'].font = title_font
sheet.row_dimensions[1].height = 22
sheet.column_dimensions['A'].width = 12
sheet.column_dimensions['E'].width = 14
sheet.column_dimensions['F'].width = 14
sheet.column_dimensions['H'].width = 20

sheet.append(excel_header)
for cell in sheet[2]:
        cell.fill = header_fill
        cell.font = header_font

index = 2
for bond_data in adjust_bonds:
    bond_code = bond_data["bond_id"]
    readjust_dt = find_property_value(adjust_bonds,bond_code, 'readjust_dt')
    #print(readjust_dt)
    if is_within_one_month(readjust_dt):
        bond_nm = find_property_value(adjust_bonds,bond_code, 'bond_nm')
        bond_id = find_property_value(adjust_bonds,bond_code, 'bond_id')
        premium_rt = find_property_value(adjust_bonds,bond_code, 'premium_rt')
        premium_rt = premium_rt/100.00
        price = find_property_value(adjust_bonds,bond_code, 'price')
        bond_backup = find_backup_content_by_id(bond_id)

        #税前收益率,需要从other bonds里面获取
        ytm_rt = find_property_value(base_bonds,bond_code, 'ytm_rt')
        #剩余到期年限
        year_left = find_property_value(base_bonds,bond_code, 'year_left')

        item = [bond_nm,int(bond_id),price,premium_rt,readjust_dt,f"{ytm_rt}%",f"{year_left}",bond_backup]
        index = index + 1
        sheet.append(item)
        #如果十天内就要满足下修条件了，那么标红
        red_font  = Font(color='FF4500')  
        if is_within_10_days(readjust_dt):
            pos = 'E'+str(index)
            sheet[pos].font = red_font


#设置百分比
for cell in sheet['D']:
    pos = 'D'+str(cell.row)
    print(sheet[pos].value)
    if(is_number(sheet[pos].value)):
        cell.number_format = '0.00%'

yahei_font  = Font(name='微软雅黑',bold=False)
for cell in sheet['A']:
    pos = 'A'+str(cell.row)
    #print(sheet[pos].value)
    #if(is_integer(sheet[pos].value)):
    cell.font = yahei_font

blue_font  = Font(color='0000FF')  # 蓝色字体
for cell in sheet['B']:
    pos = 'B'+str(cell.row)
    print(sheet[pos].value)
    if(is_integer(sheet[pos].value)):
        cell.font = blue_font

#设置备注文字样式
for cell in sheet['F']:
    pos = 'F'+str(cell.row)
    sheet[pos].font = backup_font
    #sheet[pos].alignment = Alignment(wrap_text=True)

#设置居中
# TOdo自动调整宽度
for row in sheet.iter_rows(min_row=1, max_row=index, min_col=1, max_col=8):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

list_file = get_file_path("下修重算日即将到期转债.xlsx")
workbook.save(list_file)
adjust_list_image_file = os.path.join(get_image_path(),'下修重算日即将到期转债.png')
excel2img.export_img(list_file, adjust_list_image_file, "Sheet", None)
    