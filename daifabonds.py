import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from common import *
import excel2img

request_headers_file = 'request_headers.txt'
# 读取文件内容
request_headers = read_jisilu_request_headers_file(request_headers_file)

def fetch_all_convertible_bonds():
    url = "https://www.jisilu.cn/webapi/cb/pre/?history=N"

    # 发起 HTTP GET 请求
    response = requests.get(url,headers=request_headers)

    if response.status_code == 200:
        data = response.json()
        return data.get("data", [])
    else:
        print(f"请求失败，状态码: {response.status_code}")
        return []


# 抓取所有转债数据
all_bonds_data = fetch_all_convertible_bonds()

# 创建一个新的工作簿
workbook = Workbook()
# 获取默认的工作表
# 设置头的背影颜色
header_fill = PatternFill(fill_type='solid', fgColor='f2f2f2')
sheet = workbook.active
# 创建一个字体对象并设置为默认字体
default_font = Font(name='宋体',bold=True)
# 将默认字体应用于工作表
sheet.font = default_font
sheet.default_font = default_font

sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 12
sheet.column_dimensions['H'].width = 12

# 创建字体对象
title_font = Font(name='楷体')
header_font = Font(name='宋体',bold=True)
title_height = 30

sheet.merge_cells('A1:I1')
sheet['A1'] = '待发行转债列表'
sheet['A1'].font = title_font
sheet.row_dimensions[1].height = 22

# 待发行转债列表
bond_headers = ['待发转债','转债代码','正股','正股代码','申购日期','股权登记日','百元股含权','配10张/股数','发行规模']

sheet.append(bond_headers)
#sheet.row_dimensions[2].height = 44
for cell in sheet[2]:
        cell.fill = header_fill
        cell.font = header_font

# 检查请求是否成功
tip_str = '[新债申购]:'
for bond_data in all_bonds_data:
    record_dt = bond_data["record_dt"]
    apply_date = bond_data["apply_date"]

    if extract_date_info(record_dt) != None and compare_dates(apply_date,get_today_date()):
        print(record_dt)
        bond_nm = bond_data["bond_nm"]
        bond_id = int(bond_data["bond_id"])
        stock_nm = bond_data["stock_nm"]
        stock_id = bond_data["stock_id"]
        apply_date = bond_data["apply_date"]
        record_dt = bond_data["record_dt"]
        cb_amount = bond_data["cb_amount"]
        apply10 =  bond_data["apply10"]
        amount = bond_data["amount"]
        item = [bond_nm,bond_id,stock_nm,stock_id,apply_date,record_dt,cb_amount,apply10,amount]
        sheet.append(item)
        if(apply_date == get_today_date()):
            tip_str += bond_nm


#设置居中
for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=9):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

blue_font  = Font(color='0000FF')  # 蓝色字体
for cell in sheet['B']:
    pos = 'B'+str(cell.row)
    print(sheet[pos].value)
    if(is_integer(sheet[pos].value)):
        cell.font = blue_font

for cell in sheet['D']:
    pos = 'D'+str(cell.row)
    print(sheet[pos].value)
    if(is_integer(sheet[pos].value)):
        cell.font = blue_font

list_file = get_file_path("待发行转债列表.xlsx")
workbook.save(list_file)

adjust_list_image_file = os.path.join(get_image_path(),'待发行转债列表.png')
excel2img.export_img(list_file, adjust_list_image_file, "Sheet", None)

append_reminder(tip_str)