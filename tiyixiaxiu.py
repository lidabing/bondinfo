import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from common import *
import excel2img

request_headers_file = 'request_headers.txt'
# 读取文件内容
request_headers = read_jisilu_request_headers_file(request_headers_file)

# 输出文件内容
print(request_headers)

def is_over_three_months(date_string):
    current_time = datetime.now()
    if date_string is None:
        return False

    supported_formats = ["%Y/%m/%d", "%Y-%m-%d"]

    for date_format in supported_formats:
        try:
            input_time = datetime.strptime(date_string, date_format)
            three_months_later = current_time + timedelta(days=90)
            return input_time > three_months_later
        except ValueError:
            continue

    return False
    
def is_past_time(date_string):
    current_time = datetime.now()
    if date_string is None:
        return True
    try:
        input_time = datetime.strptime(date_string, "%Y-%m-%d")
        return input_time < current_time and input_time.date() != current_time.date()
    except ValueError:
        return False

def fetch_all_convertible_bonds():
    url = "https://www.jisilu.cn/webapi/cb/adjust/"

    # 发起 HTTP GET 请求
    response = requests.get(url,headers=request_headers)

    if response.status_code == 200:
        data = response.json()
        return data.get("data", [])
    else:
        print(f"请求失败，状态码: {response.status_code}")
        return []

# 抓取所有转债数据
all_bonds_data = fetch_all_convertible_bonds()


excel_header =   ['转债名称','转债代码','收盘价','溢价率','股东大会','备注']
# 创建字体对象
title_font = Font(name='楷体')
header_font = Font(name='微软雅黑',bold=False)
backup_font = Font(name='微软雅黑',size=9)
title_height = 30
header_fill = PatternFill(fill_type='solid', fgColor='f2f2f2')


# 创建一个新的工作簿
workbook = Workbook()
sheet = workbook.active
sheet.merge_cells('A1:F1')
sheet['A1'] = '已经提议下修转债列表'
sheet['A1'].font = title_font
sheet.row_dimensions[1].height = 22
sheet.column_dimensions['A'].width = 12
sheet.column_dimensions['E'].width = 14
sheet.column_dimensions['F'].width = 14
sheet.column_dimensions['G'].width = 20

sheet.append(excel_header)
for cell in sheet[2]:
        #if(cell.internal_value != "代码"):
        cell.fill = header_fill
        cell.font = header_font
index = 2
tip_str='[今日决议下修]:'
for bond_data in all_bonds_data:
    bond_code = bond_data["bond_id"]
    adjust_count = find_property_value(all_bonds_data,bond_code, 'adjust_count')
    adjust_date = find_property_value(all_bonds_data,bond_code, 'adjust_date')
    print(adjust_date)
    if is_past_time(adjust_date):
        continue
    #print(adjust_date)
    #print(adjust_count)
    adjust = parse_adjust_string(adjust_count)
    #if adjust == None:
    #    continue
    #if adjust[0] == 0:
    #    continue
    #if adjust[0] > 15:
    #    continue
    if is_over_three_months(adjust_date):
       adjust_date = '下修时间待定'

    #再判断收盘价，如果收盘价高于122，就不统计了
    price = find_property_value(all_bonds_data,bond_code, 'price')
    if price>122:
        continue

    index = index + 1
    bond_nm = find_property_value(all_bonds_data,bond_code, 'bond_nm')
    bond_id = find_property_value(all_bonds_data,bond_code, 'bond_id')
    #bond_id = int(bond_id)
    #price = find_property_value(all_bonds_data,bond_code, 'price')
    premium_rt = find_property_value(all_bonds_data,bond_code, 'premium_rt')
    premium_rt = premium_rt/100.00
    #premium_rt_str =  str(premium_rt) + '%'
    readjust_dt = find_property_value(all_bonds_data,bond_code, 'readjust_dt')
    bond_backup = find_backup_content_by_id(bond_id)
    item = [bond_nm,int(bond_id),price,premium_rt,adjust_date,bond_backup]
    sheet.append(item)
    if(is_same_day(adjust_date)):
        tip_str +=bond_nm + ' '

    #如果只剩下3天了，那么标红，做醒目处理
    red_font  = Font(color='FF4500')  
    #if adjust[1]-adjust[0]<4:
    #    pos = 'E'+str(index)
    #    sheet[pos].font = red_font


    

#设置百分比
for cell in sheet['D']:
    pos = 'D'+str(cell.row)
    print(sheet[pos].value)
    if(is_number(sheet[pos].value)):
        cell.number_format = '0.00%'

blue_font  = Font(color='0000FF')  # 蓝色字体
for cell in sheet['B']:
    pos = 'B'+str(cell.row)
    print(sheet[pos].value)
    if(is_integer(sheet[pos].value)):
        cell.font = blue_font

yahei_font  = Font(name='微软雅黑',bold=False)
for cell in sheet['A']:
    pos = 'A'+str(cell.row)
    #print(sheet[pos].value)
    #if(is_integer(sheet[pos].value)):
    cell.font = yahei_font

#设置备注文字样式
for cell in sheet['F']:
    pos = 'F'+str(cell.row)
    sheet[pos].font = backup_font
    #sheet[pos].alignment = Alignment(wrap_text=True)
             
#设置居中
# TOdo自动调整宽度
for row in sheet.iter_rows(min_row=1, max_row=index, min_col=1, max_col=6):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
# 保存工作簿
adjust_list_file = get_file_path("已经提议下修转债列表.xlsx")
workbook.save(adjust_list_file)
adjust_list_image_file = os.path.join(get_image_path(),'已经提议下修转债列表.png')
excel2img.export_img(adjust_list_file, adjust_list_image_file, "Sheet", None)
append_reminder(tip_str)