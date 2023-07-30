#邀约股
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from common import *
import excel2img


def convert_stock_price(price_str):
    try:
        price = float(price_str.replace(',', '').replace('%', ''))
        return price
    except ValueError:
        return None
    
def convert_stock_code(stock_code):
    if stock_code.startswith('6'):
        return f"sh{stock_code}"
    else:
        return f"sz{stock_code}"

def get_stock_price(stock_code):
    url = f"https://qt.gtimg.cn/q=s_{stock_code}"
    try:
        response = requests.get(url)
        response.raise_for_status()  # 检查网络请求是否成功
        data = response.text

        # 解析数据
        fields = data.split("~")
        stock_price = float(fields[3])  # 最新股价

        return stock_price
    except (requests.exceptions.RequestException, IndexError, ValueError) as e:
        print(f"Failed to get stock price for {stock_code}: {e}")
        return None
    


excel_header =   ['代码','名称','邀约价','收盘价','阶段','备注']
# 创建字体对象
title_font = Font(name='微软雅黑')
header_font = Font(name='微软雅黑',bold=False)
backup_font = Font(name='微软雅黑',size=9)
title_height = 30
header_fill = PatternFill(fill_type='solid', fgColor='f2f2f2')


# 创建一个新的工作簿
workbook = Workbook()
sheet = workbook.active
sheet.merge_cells('A1:F1')
sheet['A1'] = '要约收购和吸收合并信息列表'
sheet['A1'].font = title_font
sheet.row_dimensions[1].height = 22
sheet.column_dimensions['A'].width = 12
sheet.column_dimensions['E'].width = 40
#sheet.column_dimensions['F'].width = 14
#sheet.column_dimensions['G'].width = 20


sheet.append(excel_header)
for cell in sheet[2]:
        cell.fill = header_fill
        cell.font = header_font

filename = "yaoyuedata.txt"  # 文件名

with open(filename, "r", encoding="utf-8") as file:
    data = file.read()  # 读取文件内容

lines = data.split('\n')  # 按行分割数据

for line in lines:
    fields = line.split('|')  # 按竖线分割字段
    if len(fields) >= 4:
        stock_code = fields[0]
        stock_name = fields[1]
        yaoyue_price = fields[2]
        yaoyue_info = fields[3]

        request_code = convert_stock_code(stock_code)
        stock_price = get_stock_price(request_code)

        item = [stock_code,stock_name,convert_stock_price(yaoyue_price),stock_price,yaoyue_info]
        sheet.append(item)

        
        print(f"股票代码: {stock_code}")
        print(f"股票名称: {stock_name}")
        print(f"要约价格: {yaoyue_price}")
        print(f"要约阶段: {yaoyue_info}")
       


blue_font  = Font(color='0000FF')  # 蓝色字体
for cell in sheet['A']:
    pos = 'A'+str(cell.row)
    print(sheet[pos].value)
    if(is_integer(sheet[pos].value)):
        cell.font = blue_font

#设置备注文字样式
for cell in sheet['E']:
    pos = 'E'+str(cell.row)
    sheet[pos].font = backup_font
    #sheet[pos].alignment = Alignment(wrap_text=True)

#设置居中
# TOdo自动调整宽度
for row in sheet.iter_rows(min_row=1, max_row=8, min_col=1, max_col=6):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)


list_file = get_file_path("要约股信息.xlsx")
workbook.save(list_file)
adjust_list_image_file = os.path.join(get_image_path(),'要约股信息.png')
excel2img.export_img(list_file, adjust_list_image_file, "Sheet", None)


    

