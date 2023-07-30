import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill


from enum import Enum
import re
from datetime import datetime
from common import *
import excel2img


# 获取一些常用属性
def fetch_all_convertible_bonds():
    request_headers_file = 'request_headers.txt'
    # 读取文件内容
    request_headers = read_jisilu_request_headers_file(request_headers_file)
    url = "https://www.jisilu.cn/webapi/cb/adjust/"

    # 发起 HTTP GET 请求
    response = requests.get(url,headers=request_headers)

    if response.status_code == 200:
        data = response.json()
        return data.get("data", [])
    else:
        print(f"请求失败，状态码: {response.status_code}")
        return []


### 通用的转债数据，在这里，主要用来获取溢价率的
all_bonds_data = fetch_all_convertible_bonds()


##################################################################
class RedeemStatus(Enum):
    NOT_REDEEM_CONDITION = 0  # 未满足强赎条件
    REDEEM_CONDITION = 1  # 满足强赎条件
    REDEEM_NOTICE = 2  # 发出强赎回公告
    NOT_REDEEM_NOTICE = 3  # 发出不强赎公告
    EXPIRING_BOND = 4  # 即将到期转债


def get_redeem_status(string):
    if "已满足强赎条件" in string:
        return RedeemStatus.REDEEM_CONDITION
    if "不强赎" in string:
        return RedeemStatus.NOT_REDEEM_NOTICE
    elif "强赎" in string:
        return RedeemStatus.REDEEM_NOTICE
    elif "到期" in string:
        return RedeemStatus.EXPIRING_BOND
    else:
        return RedeemStatus.NOT_REDEEM_CONDITION #还没有达到强赎条件


url = "https://www.jisilu.cn/webapi/cb/redeem/"

# 发送GET请求到URL
response = requests.get(url)

#转债代码 转债名称 收盘价 溢价率 剩余规模 达标天数
#转债代码  转债名称 收盘价 溢价率 最后交易日 最后转股日
#转债代码  转债名称 收盘价 溢价率 最后交易日 最后转股日
not_redeem_condition_header = ['转债名称','转债代码','收盘价','溢价率','剩余规模','达标天数']
redeem_condition_header = ['转债名称','转债代码','收盘价','溢价率','剩余规模','达标天数']
redeem_notice_header = ['转债名称','转债代码','收盘价','溢价率','最后交易日','最后转股日']
expriry_bond_header = ['转债名称','转债代码','收盘价','溢价率','最后交易日','最后转股日']

select_stocks=[]

# 检查请求是否成功
if response.status_code == 200:
    # 解析JSON响应
    json_data = response.json()
    #即将满足强赎条件的
    not_redeem_condition = [not_redeem_condition_header]
    #宣布强赎
    redeem_notice = [redeem_notice_header]
    #满足强赎条件了
    redeem_condition =  [redeem_condition_header]
    # 过期转债
    expriry_bond = [expriry_bond_header]
    #EXPIRY_BOND
    # 访问解析后的数据
    for item in json_data['data']:
        redeem_status = item["redeem_status"]
        status = get_redeem_status(redeem_status)
        premium_rt = find_property_value(all_bonds_data,item["bond_id"],'premium_rt')
        premium_rt = premium_rt/100.00
        #premium_rt = str(premium_rt) + '%'
        if status == RedeemStatus.NOT_REDEEM_CONDITION:  
            number = separate_numbers(redeem_status)
            compliance_days = int(number[0])#numbers[1]-numbers[0]
            if compliance_days>5:
                print("即将达到强赎条件")
                cache = [item["bond_nm"],int(item["bond_id"]),item["price"],premium_rt,item["curr_iss_amt"],redeem_status]
                not_redeem_condition.append(cache)            
        elif status == RedeemStatus.REDEEM_CONDITION:
            cache = [item["bond_nm"],int(item["bond_id"]),item["price"],premium_rt,item["curr_iss_amt"],redeem_status]
            redeem_condition.append(cache)

        elif status == RedeemStatus.REDEEM_NOTICE:
            print("发出强赎公告")
            cache = [item["bond_nm"],int(item["bond_id"]),item["price"],premium_rt,item["delist_dt"],item["last_convert_dt"]]
            redeem_notice.append(cache)
        elif status == RedeemStatus.EXPIRING_BOND:
            print("转债即将到期")
            cache = [item["bond_nm"],int(item["bond_id"]),item["price"],premium_rt,item["delist_dt"],item["last_convert_dt"]]
            expriry_bond.append(cache)
         

        #print("债券ID:", item["bond_id"])
        #print("债券名称:", item["bond_nm"])
        #print("赎回日期:", item["redeem_dt"])
        #print("强制赎回价格:", item["force_redeem_price"])
        #print("强制赎回价格:", item["force_redeem_price"])
        #print("-----------------------------")
    
    # 创建一个新的工作簿
    workbook = Workbook()
    # 获取默认的工作表
    # 设置头的背影颜色
    header_fill = PatternFill(fill_type='solid', fgColor='f2f2f2')
    sheet = workbook.active
    # 创建一个字体对象并设置为默认字体
    default_font = Font(name='微软雅黑',bold=False)
    # 将默认字体应用于工作表
    sheet.font = default_font
    sheet.default_font = default_font

    # 创建字体对象
    title_font = Font(name='楷体')
    header_font = Font(name='微软雅黑',bold=False)
    title_height = 30

    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['E'].width = 16
    sheet.column_dimensions['F'].width = 16

    #for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=5):
    #     for cell in row:
    #         cell.alignment = Alignment(horizontal='center', vertical='center')

    #即将达到强赎条件列表
    sheet.merge_cells('A1:F1')
    sheet['A1'] = '即将达到强赎条件列表'
    sheet.row_dimensions[1].height = 22

    #sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    # 设置单元格A1的字体样式
    sheet['A1'].font = title_font
    index = 2 #指针
    header_index = index
   
    for row in not_redeem_condition:
        sheet.append(row)
        index = index + 1
    #设置背景颜色
    for cell in sheet[header_index]:
        cell.fill = header_fill
        cell.font = header_font
        #cell.font = font
    #设置字体颜色


    size = len(redeem_condition)
    if size > 1:
        range = 'A'+str(index) +':F'+str(index)
        startpos = 'A'+str(index)
        sheet.merge_cells(range)
        sheet[startpos] = '已满足强赎条件'
        sheet.row_dimensions[index].height = title_height
        sheet[startpos].alignment = Alignment(horizontal='center', vertical='center')
        sheet[startpos].font = title_font
        index = index + 1
        header_index = index
        for row in redeem_condition:
            sheet.append(row)
            index = index + 1
        #设置背景颜色
        for cell in sheet[header_index]:
            cell.fill = header_fill
            cell.font = header_font



    #已经发出强赎公告列表
    range = 'A'+str(index) +':F'+str(index)
    startpos = 'A'+str(index)
    sheet.merge_cells(range)
    sheet[startpos] = '已经发出强赎公告列表'
    sheet.row_dimensions[index].height = title_height
    sheet[startpos].alignment = Alignment(horizontal='center', vertical='center')
    # 设置单元格A1的字体样式
    sheet[startpos].font = title_font
    index = index + 1
    header_index = index
    for row in redeem_notice:
        sheet.append(row)
        index = index + 1
    #设置背景颜色
    for cell in sheet[header_index]:
        cell.fill = header_fill
        cell.font = header_font
    #设置文字颜色

    #转债即将到期列表
    range = 'A'+str(index) +':F'+str(index)
    startpos = 'A'+str(index)
    sheet.merge_cells(range)
    sheet[startpos] = '转债即将到期列表'
    sheet.row_dimensions[index].height = title_height
    sheet[startpos].alignment = Alignment(horizontal='center', vertical='center')
    # 设置单元格的字体样式
    sheet[startpos].font = title_font
    index = index + 1
    header_index = index
    for row in expriry_bond:
        sheet.append(row)
        index = index + 1
    #设置背景颜色
    for cell in sheet[header_index]:
        #if(cell.internal_value != "代码"):
        cell.fill = header_fill
        cell.font = header_font


    blue_font  = Font(color='0000FF')  # 蓝色字体
    for cell in sheet['B']:
        #if "不强赎" in cell.internal_value:
        #    print('标题不设置蓝色')
        #else:
        pos = 'B'+str(cell.row)
        print(sheet[pos].value)
        if(is_integer(sheet[pos].value)):
            cell.font = blue_font

    #设置百分比
    for cell in sheet['D']:
        pos = 'D'+str(cell.row)
        print(sheet[pos].value)
        if(is_number(sheet[pos].value)):
            cell.number_format = '0.00%'


    #设置居中
    for row in sheet.iter_rows(min_row=1, max_row=15, min_col=1, max_col=6):
         for cell in row:
             cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
             
         

    # 保存工作簿
    list_file = get_file_path("强赎转债列表.xlsx")
    workbook.save(list_file)
    adjust_list_image_file = os.path.join(get_image_path(),'强赎转债列表.png')
    excel2img.export_img(list_file, adjust_list_image_file, "Sheet", None)

else:
    print("请求失败，状态码为:", response.status_code)


